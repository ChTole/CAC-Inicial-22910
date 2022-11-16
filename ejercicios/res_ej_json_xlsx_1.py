"""
Dificultad adicional 1: generar un nuevo script de Python que realice una búsqueda dentro del libro creado permitiendo seleccionar la columna a filtrar.
"""

from openpyxl import load_workbook

def cargar_libro(ruta):
    try:
        libro = load_workbook(filename= ruta)
    except FileNotFoundError:
        print('No encontré el archivo!')
        ruta = input('Ingrese ubicación correcta del archivo: ')
        libro = load_workbook(filename= ruta)
    return libro

mi_libro = cargar_libro(r'D:\CodoACodo\Comision 22910\Clase30\Datos.xlsx')  

def buscar_en_hoja(columna, busqueda):
    resultado = False
    columna = int(columna)
    hoja = mi_libro.active
    for fila in hoja.values:
        if busqueda.lower() in str(fila[columna]).lower():
            print(f'Encontré {fila[0]} // {fila[1]} - {fila[2]} {fila[3]} {fila[4]}')
            resultado = True
    if not resultado:
        print('No encontré lo que buscabas!')
        
            

columna = ''
while columna != '5':
    print("""
        Opciones de búsqueda:
        0 - Por Id
        1 - Por Nombre
        2 - Por Correo
        3 - Por Teléfono
        4 - Por Sitio web
        5 - Salir del menú          
            """)
    columna = input('Ingrese columna de búsqueda: ')
    if int(columna) in range(5):
        busqueda = input('Ingrese valor a buscar: ')
        buscar_en_hoja(columna, busqueda)
    elif columna == '5':
        print('Gracias por usar la app!')
        continue
    else:
        print('Opción inválida!')
        columna = input('Ingrese columna de búsqueda: ')




# Auxiliar para ver los tipos de datos en la hoja
# for fila in mi_libro.active.values:
#     for celda in fila:
#         print(type(celda))
        
 
