"""
Dificultad adicional 2: generar un nuevo script de Python que permita agregar una columna llamada "fecha de nacimiento" y solicite por consola el ingreso de los datos para cada usuario. Utilizar el módulo datetime y corroborar que cada dato ingresado para construir la fecha sea de acuerdo al formato dd/mm/aaaa. Por ejemplo:
- si ingreso 23/04/1998, la fecha se crea
- si ingreso 43/15/2005, la fecha no se crea.
"""

from openpyxl import load_workbook
from datetime import date

def cargar_libro(ruta):
    try:
        libro = load_workbook(filename= ruta)
    except FileNotFoundError:
        print('No encontré el archivo!')
        ruta = input('Ingrese ubicación correcta del archivo: ')
        libro = load_workbook(filename= ruta)
    return libro

mi_libro = cargar_libro(r'D:\CodoACodo\Comision 22910\Clase31\Datos.xlsx')  

hoja = mi_libro.active
hoja["f1"] = "Fecha de nacimiento"

for fila in hoja.iter_rows(min_row=2, max_row=hoja.max_row):
    fecha_nac = input(f'Ingrese fecha de nacimiento para {fila[1].value} (dd/mm/aaaa): ')
    while True:
        try:
            fila[5].value = date(
                int(fecha_nac.split('/')[2]),
                int(fecha_nac.split('/')[1]),
                int(fecha_nac.split('/')[0]),
            )
            break
        except Exception as e:
            print(type(e).__name__) # para ver cual fue la excepción
            print('Fecha inválida!')
            fecha_nac = input(f'Ingrese fecha de nacimiento para {fila[1].value} (dd/mm/aaaa): ')
        
    # date(1981, 10, 15)
    # fila[5].value = fecha_nac

while True:
    try:
        mi_libro.save(filename=r'D:\CodoACodo\Comision 22910\Clase31\Datos.xlsx')
        mi_libro.close()
        break
    except PermissionError:
        print('El libro está abierto, por favor cerralo para implementar los cambios!!!')
        input('Presiona ENTER luego de cerrarlo >>>')

print('Cambios guardados exitosamente!')